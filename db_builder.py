import sqlite3
import openpyxl

# create table
def create_db():
    """
    create mysql database "SampleSuperStore.db" with features
    as column names from existing xlsx file    
    """
    # connect to mysql
    conn = sqlite3.connect("SSS.db")

    # create a cursor object to create table
    cursor=conn.cursor()

    # create the table
    table_cols = '''
        create table if not exists SampleSuperStore(
                   row_ID int, 
                   order_ID varchar(25),
                   order_date date,
                   ship_date date,
                   ship_mode varchar(25),
                   customer_ID varchar(25),
                   customer_name varchar(25),
                   segment varchar(25),
                   Country_Region varchar(25),
                   City varchar(25),
                   State varchar(25),
                   PostCode int,
                   Region varchar(25),
                   product_ID varchar(25),
                   category varchar(25),
                   sub_category varchar(25),
                   product_name varchar(25),
                   sales float,
                   quantity int,
                   discount float,
                   profit float,
                   primary key(row_ID)
                   );
                '''
    cursor.execute(table_cols)

    # insert table values
    
    # load excel file
    excel = openpyxl.load_workbook('SampleSuperStore.xlsx')
    excel = excel[excel.sheetnames[0]]
    
    for _, row in enumerate(excel.iter_rows(min_row=2)):
        # get the values from each cell
        row_ID=row[0].value
        order_ID=row[1].value
        order_date=row[2].value
        ship_date=row[3].value
        ship_mode=row[4].value
        customer_ID=row[5].value
        customer_name=row[6].value
        segment=row[7].value
        Country_Region=row[8].value
        City=row[9].value
        State=row[10].value
        PostCode=row[11].value
        Region=row[12].value
        product_ID=row[13].value
        category=row[14].value
        sub_category=row[15].value
        product_name=row[16].value
        sales=row[17].value
        quantity=row[18].value
        discount=row[19].value
        profit=row[20].value

        # insert values
        cursor.execute('''
        INSERT OR REPLACE INTO SampleSuperStore 
        (row_ID,order_ID,order_date,ship_date,ship_mode,customer_ID,
        customer_name,segment,Country_Region,City,State,PostCode,
        Region,product_ID,category,sub_category,product_name,sales,
        quantity,discount,profit)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ''',(row_ID,order_ID,order_date,ship_date,ship_mode,customer_ID,
        customer_name,segment,Country_Region,City,State,PostCode,
        Region,product_ID,category,sub_category,product_name,sales,
        quantity,discount,profit))

    # commit the changes
    conn.commit()
        
    #close the connection
    conn.close()
